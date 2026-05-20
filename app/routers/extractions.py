# app/routers/extractions.py
import csv
import io
import json
import logging
from typing import List, Optional
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc

from app.db.connection import get_db
from app.db.models import Extraction, PromptConfig
from app.schemas.extraction import ExtractionRead, ExtractionValidateRequest
from app.services.response_validator import validate_and_clean_response, ResponseValidationError

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/extractions", tags=["extractions"])


@router.get("/", response_model=List[ExtractionRead])
async def list_extractions(
    config_id: Optional[UUID] = Query(None, description="Filtrar por config_id"),
    status: Optional[str] = Query(None, description="Filtrar por estado: success | failed | validated"),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """
    Devuelve el historial de extracciones, ordenado de más reciente a más antigua.
    Permite filtrar por config_id y/o status.
    """
    try:
        stmt = select(Extraction).order_by(desc(Extraction.created_at)).limit(limit)

        if config_id is not None:
            stmt = stmt.where(Extraction.prompt_config_id == config_id)

        if status is not None:
            stmt = stmt.where(Extraction.status == status)

        result = await db.execute(stmt)
        return result.scalars().all()

    except Exception as exc:
        logger.exception("Error al listar extracciones")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(exc)}")


@router.get("/{extraction_id}", response_model=ExtractionRead)
async def get_extraction(
    extraction_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """
    Devuelve el detalle completo de una extracción, incluyendo el prompt enviado
    y la respuesta cruda del LLM.
    """
    try:
        result = await db.execute(
            select(Extraction).where(Extraction.id == extraction_id)
        )
        extraction = result.scalars().first()

        if not extraction:
            raise HTTPException(status_code=404, detail="Extracción no encontrada")

        return extraction

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al obtener extracción")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(exc)}")


@router.patch("/{extraction_id}/validate", response_model=ExtractionRead)
async def validate_extraction(
    extraction_id: UUID,
    payload: ExtractionValidateRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Guarda la validación humana de una extracción y la marca como validated.
    """
    try:
        extraction_result = await db.execute(
            select(Extraction).where(Extraction.id == extraction_id)
        )
        extraction = extraction_result.scalars().first()

        if not extraction:
            raise HTTPException(status_code=404, detail="Extracción no encontrada")

        config_result = await db.execute(
            select(PromptConfig).where(PromptConfig.id == extraction.prompt_config_id)
        )
        config = config_result.scalars().first()

        if not config:
            raise HTTPException(status_code=404, detail="Configuración de prompt no encontrada")

        try:
            cleaned = validate_and_clean_response(payload.result, config.variables)
        except ResponseValidationError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        extraction.validated_result = cleaned
        extraction.status = "validated"
        extraction.error_message = None

        db.add(extraction)
        await db.commit()
        await db.refresh(extraction)

        return extraction

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al validar extracción")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(exc)}")


def _get_result_rows(extraction: Extraction) -> list[dict]:
    """Devuelve la lista [{title, answer}] priorizing validated_result sobre raw_llm_response."""
    if extraction.validated_result:
        return extraction.validated_result
    if extraction.raw_llm_response:
        try:
            parsed = json.loads(extraction.raw_llm_response)
            if isinstance(parsed, list):
                return parsed
        except (json.JSONDecodeError, TypeError):
            pass
    return []


@router.get("/{extraction_id}/export/xlsx")
async def export_extraction_xlsx(
    extraction_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """Exporta una extracción individual como archivo Excel (.xlsx)."""
    result = await db.execute(select(Extraction).where(Extraction.id == extraction_id))
    extraction = result.scalars().first()
    if not extraction:
        raise HTTPException(status_code=404, detail="Extracción no encontrada")

    rows = _get_result_rows(extraction)
    doc_name = extraction.document_name or str(extraction.id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracción"

    # Cabecera
    ws.append(["Campo", "Valor"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    for row in rows:
        ws.append([row.get("title", ""), row.get("answer", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_." else "_" for c in doc_name)
    filename = f"extraccion_{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/bulk")
async def export_extractions_bulk(
    format: str = Query("csv", description="Formato: csv | xlsx | json"),
    config_id: Optional[UUID] = Query(None),
    collection_id: Optional[UUID] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
):
    """
    Exporta múltiples extracciones en formato CSV, XLSX o JSON.
    Las columnas se construyen dinámicamente a partir de los campos extraídos.
    Cada fila = un documento; cada columna = un campo extraído.
    """
    if format not in ("csv", "xlsx", "json"):
        raise HTTPException(status_code=400, detail="Formato debe ser csv, xlsx o json")

    stmt = select(Extraction).order_by(desc(Extraction.created_at)).limit(limit)
    if config_id is not None:
        stmt = stmt.where(Extraction.prompt_config_id == config_id)
    if collection_id is not None:
        stmt = stmt.where(Extraction.collection_id == collection_id)
    if status is not None:
        stmt = stmt.where(Extraction.status == status)

    result = await db.execute(stmt)
    extractions = result.scalars().all()

    # Recopilar todas las columnas de campos en orden de aparición
    field_cols: list[str] = []
    seen: set[str] = set()
    records = []
    for ext in extractions:
        rows = _get_result_rows(ext)
        rec = {
            "id": str(ext.id),
            "documento": ext.document_name or "",
            "estado": ext.status,
            "config_id": str(ext.prompt_config_id),
            "latencia_ms": ext.latency_ms,
            "fecha": ext.created_at.isoformat() if ext.created_at else "",
        }
        for row in rows:
            t = row.get("title", "")
            rec[t] = row.get("answer", "")
            if t and t not in seen:
                seen.add(t)
                field_cols.append(t)
        records.append(rec)

    meta_cols = ["id", "documento", "estado", "config_id", "latencia_ms", "fecha"]
    all_cols = meta_cols + field_cols

    if format == "json":
        content = json.dumps(records, ensure_ascii=False, indent=2)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/json",
            headers={"Content-Disposition": 'attachment; filename="extracciones.json"'},
        )

    if format == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=all_cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),  # utf-8-sig para Excel
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="extracciones.csv"'},
        )

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracciones"
    ws.append(all_cols)
    for rec in records:
        ws.append([rec.get(c, "") for c in all_cols])
    for i, col in enumerate(all_cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(15, len(col) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="extracciones.xlsx"'},
    )
